import sqlite3
from openpyxl import load_workbook
import sys
import numpy as np
import math

connection = sqlite3.connect("db.sqlite3")
crsr = connection.cursor()

sql_command = "SELECT userID FROM users"
crsr.execute(sql_command)
users = crsr.fetchall()

factors = ['A','B','C','E','F','G','H','I','L','M','N','O','P','Q1','Q2','Q3','Q4']
scales = ['HAV', 'HAA', 'HAD', 'HFV', 'HFA', 'HFD', 'PAV', 'PAA', 'PAD', 'PSP', 'PSE', 'PSS', 'PSA', 'PSS1']
i = 0
for factor in factors:
    i += 1
    j = 0
    for scale in scales:
        j += 1
        first_array = []
        second_array = []
        for user in users:
            sql_command = "SELECT point FROM results WHERE userID = " + user + "AND factor = '" + factor + "'"
            crsr.execute(sql_command)
            factor_point = crsr.fetchall()[0][0]
            sql_command = "SELECT point FROM logs WHERE userID = " + user + "AND scale = '" + scale + "'"
            crsr.execute(sql_command)
            scale_point = crsr.fetchall()[0][0]
            first_array.append(factor_point)
            second_array.append(scale_point)
        wb = load_workbook("analysis.xlsx")
        coef = np.corrcoef(first_array, second_array)
        wb['corr'].cell(row=i,column=j).value = coef
        wb['corr'].cell(row=j,column=i).value = coef